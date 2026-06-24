"""
Importa la LISTA DE PRECIOS.xlsx a un catálogo limpio (lib/data/catalog.json).

Reglas:
- Columna B = código, C = nombre, D = peso (kg), H = PRECIO PÚBLICO.
- NUNCA exportamos la columna E (costo interno) ni F (liquidación).
- Las categorías van como filas-título en B; se descartan frases religiosas
  y separadores. Las secciones sin productos se eliminan solas.
- Precios #VALUE!/vacíos -> null (se mostrará "Cotizar").

Uso:  python scripts/import-catalog.py
"""
import json
import re
import unicodedata
from datetime import date
from pathlib import Path

import openpyxl

# Ruta al Excel original del cliente (fuera del repo).
SRC = Path(
    r"C:\Users\PERSONAL\Documents\MULTIACEROS B&M S.A.S"
    r"\MULTIACEROS B&M S.A.S\INFORMACION DE LA EMPRESA\LISTA DE PRECIOS.xlsx"
)
OUT = Path(__file__).resolve().parent.parent / "lib" / "data" / "catalog.json"

CODE = re.compile(r"^\*?\d{4,8}$")
RELIGIOUS = (
    "DIOS", "SENOR", "JEHOV", "CRISTO", "JESUC", "MISERICORD", "GRACIA",
    "BENDITO", "SALVAC", "OMNIPOTENTE", "REFUGIO", "ANGUSTIA", "TRONO",
    "SOCORRO", "SOMBRA", "ALAS", "SENDAS", "FORTALEZA", "VICTORIA",
    "POTENTE", "LAMENTO", "BAILE", "CONFIA", "BRONCE", "TESOROS", "PORCION",
    "VERDAD Y", "ABORRECED", "RECTITUD", "EFICAZ", "PALABRA", "REFUERZA",
    "QUEBRANTA", "ENDEREZAR", "TORCIDOS", "PEDAZOS", "ISRAEL", "ESPERARE",
    "LUZ Y", "CERROJOS DE", "PUERTAS DE",
)

FAMILIES = [
    ("tuberia-perfileria", "Tubería y Perfilería", "/images/tuberia-perfiles.jpg"),
    ("laminas-aceros", "Láminas y Aceros", "/images/aceros-variedades.jpg"),
    ("angulos-platinas-varilla", "Ángulos, Platinas y Varilla", "/images/angulos.jpg"),
    ("cubiertas-tejas", "Cubiertas y Tejas", "/images/cubiertas-arquitectonicas.jpg"),
    ("cemento-pvc-obra-gris", "Cemento, PVC y Obra Gris", "/images/cemento-oriente.jpg"),
    ("cerramiento-alambre-malla", "Cerramiento, Alambre y Malla", "/images/ladrillos.jpg"),
    ("pintura-soldadura-quimicos", "Pintura, Soldadura y Químicos", "/images/teja-colonial.jpg"),
    ("herramientas-ferreteria", "Herramientas y Ferretería", "/images/kit-herramientas.jpg"),
]


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def norm(s: str) -> str:
    return strip_accents(str(s)).upper().strip()


def is_religious(s: str) -> bool:
    u = norm(s)
    return any(k in u for k in RELIGIOUS)


def family_of(section: str) -> str:
    s = norm(section)
    has = lambda *ks: any(k in s for k in ks)
    if has("ANGULO", "PLATINA", "VARILLA"):
        return "angulos-platinas-varilla"
    if has("LAMINA", "HOJALATA", "CANAL"):
        return "laminas-aceros"
    if has("TEJA", "CUBIERTA", "TERMOACUSTICA", "CABALLETE", "FLANCHE",
           "AMARRE", "PUNTA COPA", "SKY", "ZINC TERMO"):
        return "cubiertas-tejas"
    if has("TUBERIA", "PERLIN", "PERFIL", "VIGA", "FLEJE"):
        return "tuberia-perfileria"
    if has("ALAMBRE", "CONCERTINA", "GAVION", "MALLA"):
        return "cerramiento-alambre-malla"
    if has("CEMENTO", "PVC", "LADRILLO"):
        return "cemento-pvc-obra-gris"
    if has("ANTICORROSIVO", "ESMALTE", "AEROSOL", "THINNER", "MASILLA",
           "HUESO DURO", "SOLDADURA", "LIJA", "PINTURA"):
        return "pintura-soldadura-quimicos"
    return "herramientas-ferreteria"


def slugify(s: str) -> str:
    s = strip_accents(str(s)).lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:60] or "item"


def parse_price(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(round(v)) if v and v > 0 else None
    txt = str(v).replace(".", "").replace(",", "").replace("$", "").strip()
    return int(txt) if txt.isdigit() and int(txt) > 0 else None


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Hoja1"]

    section = None
    products = []
    phrases = []
    seen_slugs = {}

    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        h = ws.cell(r, 8).value
        if b is None:
            continue
        bs = str(b).strip()
        if not bs:
            continue

        if CODE.match(bs):  # producto
            name = str(c).strip() if c else ""
            if not name or norm(name) in ("CODIGO", "NOMBRE") or section is None:
                continue
            weight = d if isinstance(d, (int, float)) else None
            base = slugify(name)
            slug = base
            if slug in seen_slugs:
                seen_slugs[base] += 1
                slug = f"{base}-{seen_slugs[base]}"
            else:
                seen_slugs[base] = 0
            products.append({
                "code": bs.lstrip("*"),
                "name": name,
                "slug": slug,
                "price": parse_price(h),
                "weightKg": round(weight, 2) if weight else None,
                "section": section,
                "family": family_of(section),
            })
        else:  # posible título de sección o ruido
            up = norm(bs)
            if up in ("CODIGO", "FECHA") or up.startswith("FECHA"):
                continue
            if is_religious(bs):
                # Se CONSERVAN en los datos (no se muestran en el sitio).
                phrases.append({"text": bs, "section": section})
                continue
            if len(bs) > 45:
                continue
            section = bs  # nueva sección (las vacías se eliminan luego)

    # Conteos por sección y familia (solo secciones con productos)
    sec_counts = {}
    fam_counts = {f[0]: 0 for f in FAMILIES}
    for p in products:
        sec_counts[p["section"]] = sec_counts.get(p["section"], 0) + 1
        fam_counts[p["family"]] += 1

    families = [
        {"slug": s, "name": n, "image": img, "count": fam_counts[s]}
        for (s, n, img) in FAMILIES
    ]
    sections = [
        {"name": name, "family": family_of(name), "count": cnt}
        for name, cnt in sorted(sec_counts.items())
    ]

    data = {
        "generatedAt": date.today().isoformat(),
        "source": "LISTA DE PRECIOS.xlsx",
        "totals": {
            "products": len(products),
            "withPrice": sum(1 for p in products if p["price"]),
            "sections": len(sections),
        },
        "families": families,
        "sections": sections,
        "products": products,
        "phrases": phrases,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=0), encoding="utf-8")

    # Resumen
    print(f"Productos: {len(products)}  |  con precio: {data['totals']['withPrice']}"
          f"  |  secciones: {len(sections)}  |  frases conservadas: {len(phrases)}")
    print("Por familia:")
    for f in families:
        print(f"  - {f['name']:<32} {f['count']}")
    print(f"\nJSON escrito en: {OUT}")
    print("\nMuestra de secciones detectadas:")
    for s in sections[:12]:
        print(f"  · {s['name']} ({s['count']})  -> {s['family']}")


if __name__ == "__main__":
    main()
